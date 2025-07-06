import pandas as pd
from openpyxl.utils import get_column_letter
from .models import App, Comment, User

def export_all_to_excel(file_path: str ="exported_data.xlsx") -> None:
    """
    Export all App, Comment, and User data from the database into an Excel file.

    The Excel file will have three sheets: 'Apps', 'Comments', and 'Users'.

    Args:
        file_path (str): The path where the Excel file will be saved. Defaults to "exported_data.xlsx".

    Returns:
        None
    """
    # export Apps
    apps = App.objects.all().values(
        "id", "name", "description", "installs", "size", "updated_at", "image_urls"
    )
    apps_df = pd.DataFrame(apps)

    if not apps_df.empty and 'updated_at' in apps_df.columns:
        apps_df['updated_at'] = pd.to_datetime(apps_df['updated_at']).dt.date

    # export Comments
    comments = Comment.objects.select_related("app", "user").all().values(
        "id", "app__id", "user__id", "text", "rating", "comment_date"
    )
    comments_df = pd.DataFrame(comments).rename(columns={
        "app__id": "app_id",
        "user__id": "user_id",
    })

    if not comments_df.empty and 'comment_date' in comments_df.columns:
        comments_df['comment_date'] = pd.to_datetime(comments_df['comment_date']).dt.date

    # export Users
    users = User.objects.all().values("id", "user_id", "display_name")
    users_df = pd.DataFrame(users)

    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        apps_df.to_excel(writer, sheet_name="Apps", index=False)
        comments_df.to_excel(writer, sheet_name="Comments", index=False)
        users_df.to_excel(writer, sheet_name="Users", index=False)

        workbook = writer.book

        # format 'updated_at' column in Apps sheet as yyyy-mm-dd
        if not apps_df.empty and 'updated_at' in apps_df.columns:
            sheet = workbook["Apps"]
            col_idx = apps_df.columns.get_loc("updated_at") + 1  # 1-based index
            col_letter = get_column_letter(col_idx)
            for cell in sheet[f"{col_letter}2":f"{col_letter}{sheet.max_row}"]:
                for c in cell:
                    c.number_format = "yyyy-mm-dd"

        # format 'comment_date' column in Comments sheet as yyyy-mm-dd
        if not comments_df.empty and 'comment_date' in comments_df.columns:
            sheet = workbook["Comments"]
            col_idx = comments_df.columns.get_loc("comment_date") + 1
            col_letter = get_column_letter(col_idx)
            for cell in sheet[f"{col_letter}2":f"{col_letter}{sheet.max_row}"]:
                for c in cell:
                    c.number_format = "yyyy-mm-dd"

    print(f"Export completed, file path: {file_path}")

def run_exporter() -> None:
    """
    Runs the export_all_to_excel function to export data to the default Excel file.

    Returns:
        None
    """
    export_all_to_excel()
